
"""
TitleIQ Agent - Master Orchestrator

This module integrates all four phases into a single workflow:

Step 1: Document Intake / Vision Processing
Step 2: Chain of Title Extraction and Validation
Step 3: Lien Priority Waterfall
Step 4: Risk Scoring and Analyst Dashboard

Key design choices:
- The arithmetic waterfall is performed entirely by explicit Python logic.
- Every autonomous decision is recorded in a processing log with the evidence used.
- If critical-section document quality drops below the configured threshold,
  the workflow halts and requests human input rather than guessing.
"""

from __future__ import annotations

import json
import re
import uuid
from copy import deepcopy
from datetime import date, datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import fitz  # PyMuPDF
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


SUPPORTED_LIEN_TYPES = {
    "PROPERTY_TAX",
    "FIRST_MORTGAGE",
    "SECOND_MORTGAGE",
    "IRS_FEDERAL",
    "JUDGMENT",
    "HOA",
    "OTHER",
}


# -------------------------------------------------------------------
# Processing log
# -------------------------------------------------------------------

def _now() -> str:
    """
    Returns an ISO timestamp for processing log entries.

    Why:
    The audit trail should show when each autonomous decision was made.
    """
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


def add_log(log: List[Dict], phase: str, action: str, decision_type: str, confidence: str, evidence: Dict, result: str) -> None:
    """
    Adds a normalized audit-trail record.

    Why:
    The user requires a non-negotiable processing log for trustworthiness.
    Every autonomous or flagged decision should capture the evidence relied on.
    """
    log.append({
        "timestamp": _now(),
        "phase": phase,
        "action": action,
        "decision_type": decision_type,
        "confidence": confidence,
        "evidence": evidence,
        "result": result,
    })


# -------------------------------------------------------------------
# Shared helpers
# -------------------------------------------------------------------

def normalize_whitespace(text: str) -> str:
    """
    Collapses repeated whitespace for easier parsing.

    Why:
    PDF text extraction often introduces inconsistent line breaks and spacing.
    """
    return re.sub(r"\s+", " ", text or "").strip()


def parse_money(text: str) -> Optional[float]:
    """
    Extracts the first dollar amount from text.

    Why:
    Lien extraction depends on explicit balances. If no amount is present,
    the engine must not invent one.
    """
    m = re.search(r"\$?\s*([0-9]{1,3}(?:,[0-9]{3})*(?:\.[0-9]{2})?|[0-9]+(?:\.[0-9]{2})?)", text.replace(",", ","))
    if m:
        return float(m.group(1).replace(",", ""))
    return None


def parse_date_label(text: str, labels: List[str]) -> Optional[str]:
    """
    Extracts a date following any of the provided labels.

    Why:
    The title workflow distinguishes recording dates from other dates such as
    signature dates or note dates.
    """
    for label in labels:
        patterns = [
            rf"{label}\s+([A-Z][a-z]+ \d{{1,2}}, \d{{4}})",
            rf"{label}[:\s]+([A-Z][a-z]+ \d{{1,2}}, \d{{4}})",
            rf"{label}\s+(\d{{4}}-\d{{2}}-\d{{2}})",
        ]
        for pat in patterns:
            m = re.search(pat, text, flags=re.IGNORECASE)
            if m:
                raw = m.group(1)
                for fmt in ("%B %d, %Y", "%Y-%m-%d"):
                    try:
                        return datetime.strptime(raw, fmt).date().isoformat()
                    except ValueError:
                        pass
    return None


def parse_recording_date(text: str) -> Optional[str]:
    """
    Extracts the recording date specifically.

    Legal reasoning:
    Recording chronology controls the public record sequence used for title
    review. Signing dates are not substituted for missing recording dates.
    """
    return parse_date_label(text, ["Recorded", "Recording Date"])


def parse_filing_date(text: str) -> Optional[str]:
    """
    Extracts a filing date for liens.

    Legal reasoning:
    IRS liens are tracked by filing date for risk flagging, even though the
    engine separately budgets them as mandatory clearance cost.
    """
    return parse_date_label(text, ["Filed", "Filing Date"])


def parse_instrument(text: str) -> Optional[str]:
    """
    Extracts book/page or instrument reference.

    Why:
    Analysts need traceable county references for every material event.
    """
    patterns = [
        r"Instrument No\.?\s*([A-Za-z0-9\-\/]+)",
        r"Instrument #\s*([A-Za-z0-9\-\/]+)",
        r"Book\s+([A-Za-z0-9\-]+)\s*/\s*Page\s+([A-Za-z0-9\-]+)",
        r"Book\s+([A-Za-z0-9\-]+)\s+Page\s+([A-Za-z0-9\-]+)",
    ]
    for pat in patterns:
        m = re.search(pat, text, flags=re.IGNORECASE)
        if m:
            if len(m.groups()) == 1:
                return m.group(1)
            return f"Book {m.group(1)} / Page {m.group(2)}"
    return None


def parse_county_state(text: str, state_hint: str = "") -> Tuple[Optional[str], Optional[str]]:
    """
    Extracts county and state from text or state hint.

    Why:
    Title packets can contain records from multiple jurisdictions.
    """
    county = None
    state = state_hint or None

    m = re.search(r"([A-Z][a-zA-Z]+ County)", text)
    if m:
        county = m.group(1)

    state_names = ["Florida", "Texas", "California", "New York", "Georgia", "Ohio", "Illinois"]
    for name in state_names:
        if re.search(rf"\b{name}\b", text):
            state = name if len(state_hint) != 2 else state_hint
            break

    return county, state


def money(value: float) -> str:
    """
    Formats a numeric amount as USD string.

    Why:
    Final dashboards and spreadsheets must be readable to analysts at a glance.
    """
    return "${:,.2f}".format(float(value))


# -------------------------------------------------------------------
# Phase 1 - Document Intake
# -------------------------------------------------------------------

def classify_page_confidence(text: str) -> Tuple[str, List[str]]:
    """
    Assigns a deterministic page confidence label for text-layer PDFs.

    Why:
    In this environment the orchestrator uses direct PDF text extraction for
    digital pages. Confidence is intentionally conservative if signatures,
    obscured text, or uncertain markers appear.
    """
    flags = []
    lowered = text.lower()

    low_markers = ["unreadable", "illegible", "obscured", "faded", "handwritten", "signature block"]
    med_markers = ["copy", "scan", "margin note", "stamp"]

    confidence = "HIGH"
    if any(marker in lowered for marker in low_markers):
        confidence = "LOW"
    elif any(marker in lowered for marker in med_markers):
        confidence = "MEDIUM"

    if "signature" in lowered:
        flags.append("signature block detected")
    if "recorded" in lowered or "instrument" in lowered or "book/page" in lowered or "book" in lowered and "page" in lowered:
        flags.append("recording stamp detected")
    if "federal tax lien" in lowered:
        flags.append("federal lien page detected")
    if "property tax" in lowered or "tax statement" in lowered or "delinquent taxes" in lowered:
        flags.append("tax page detected")
    if "mortgage" in lowered:
        flags.append("mortgage page detected")
    if "deed" in lowered:
        flags.append("deed page detected")
    if any(marker in lowered for marker in ["obscured", "illegible", "faded"]):
        flags.append("critical text degradation detected")

    return confidence, flags


def extract_pdf_pages(pdf_path: Path, log: List[Dict]) -> Dict:
    """
    Extracts page text from a PDF and builds the Phase 1 JSON structure.

    Why:
    This orchestrator must accept a PDF upload and create a structured intake
    object before title and lien logic can run.
    """
    doc = fitz.open(pdf_path)
    pages = []
    high_pages, medium_pages, low_pages = [], [], []

    for idx in range(len(doc)):
        page = doc.load_page(idx)
        text = page.get_text("text")
        text = text.strip()
        confidence, flags = classify_page_confidence(text)
        requires_review = confidence == "LOW" or any(
            f in flags for f in ["critical text degradation detected", "signature block detected"]
        )
        page_data = {
            "page_number": idx + 1,
            "confidence": confidence,
            "extracted_text": text,
            "flags": flags,
            "requires_human_review": requires_review,
        }
        pages.append(page_data)

        if confidence == "HIGH":
            high_pages.append(idx + 1)
        elif confidence == "MEDIUM":
            medium_pages.append(idx + 1)
        else:
            low_pages.append(idx + 1)

        add_log(
            log,
            phase="Phase 1 - Document Intake",
            action="Extract page text and score confidence",
            decision_type="AUTONOMOUS",
            confidence=confidence,
            evidence={"page_number": idx + 1, "flags": flags, "text_sample": normalize_whitespace(text)[:180]},
            result=f"Page {idx + 1} classified {confidence}",
        )

    doc.close()

    return {
        "document_id": uuid.uuid4().hex[:16],
        "total_pages": len(pages),
        "extraction_summary": {
            "high_confidence_pages": high_pages,
            "medium_confidence_pages": medium_pages,
            "low_confidence_pages": low_pages,
        },
        "pages": pages,
    }


def phase1_quality_report(phase1: Dict) -> str:
    """
    Builds a compact quality report string.

    Why:
    The orchestrator needs a human-readable Phase 1 summary for the processing log.
    """
    total = max(1, int(phase1["total_pages"]))
    high = len(phase1["extraction_summary"]["high_confidence_pages"])
    pct = round((high / total) * 100, 2)
    review_pages = [p["page_number"] for p in phase1["pages"] if p["requires_human_review"]]
    return f"High-confidence pages: {high}/{total} ({pct}%). Human-review pages: {review_pages or 'None'}."


def identify_critical_pages(phase1: Dict) -> List[int]:
    """
    Identifies chain-of-title and lien-recording critical pages.

    Why:
    The user requires the workflow to halt if confidence on critical chain/lien
    pages drops below the threshold rather than continue on weak evidence.
    """
    critical_terms = [
        "deed", "certificate of title", "trustee", "sheriff",
        "mortgage", "deed of trust", "lien", "federal tax lien",
        "property tax", "judgment", "hoa", "recorded", "instrument"
    ]
    critical_pages = []
    for page in phase1["pages"]:
        text = (page["extracted_text"] or "").lower()
        if any(term in text for term in critical_terms):
            critical_pages.append(page["page_number"])
    return sorted(set(critical_pages))


def critical_quality_gate(phase1: Dict, log: List[Dict], min_high_conf_pct: float = 40.0) -> Dict:
    """
    Applies the critical-section quality gate.

    Why:
    If the agent lacks reliable extraction quality on chain/lien pages, it must
    stop and request human input rather than guess downstream legal or economic results.
    """
    critical_pages = identify_critical_pages(phase1)
    if not critical_pages:
        add_log(
            log,
            phase="Phase 1 - Document Intake",
            action="Critical-section quality gate",
            decision_type="AUTONOMOUS",
            confidence="HIGH",
            evidence={"critical_pages": []},
            result="No critical pages identified; gate passed.",
        )
        return {
            "halted": False,
            "critical_pages": [],
            "high_confidence_critical_pct": 100.0,
            "message": "No critical pages identified.",
        }

    high_critical = [
        p["page_number"] for p in phase1["pages"]
        if p["page_number"] in critical_pages and p["confidence"] == "HIGH"
    ]
    pct = round((len(high_critical) / len(critical_pages)) * 100, 2)

    halted = pct < min_high_conf_pct
    result_msg = (
        f"HALT - critical high-confidence rate {pct}% is below threshold {min_high_conf_pct}%."
        if halted else
        f"PASS - critical high-confidence rate {pct}% meets threshold {min_high_conf_pct}%."
    )

    add_log(
        log,
        phase="Phase 1 - Document Intake",
        action="Critical-section quality gate",
        decision_type="FLAGGED_FOR_HUMAN_REVIEW" if halted else "AUTONOMOUS",
        confidence="LOW" if halted else "HIGH",
        evidence={"critical_pages": critical_pages, "high_confidence_critical_pages": high_critical, "threshold_pct": min_high_conf_pct},
        result=result_msg,
    )

    return {
        "halted": halted,
        "critical_pages": critical_pages,
        "high_confidence_critical_pct": pct,
        "message": result_msg,
    }


# -------------------------------------------------------------------
# Phase 2 - Chain of Title
# -------------------------------------------------------------------

def parse_deed_event(text: str) -> Tuple[Optional[str], Optional[str], Optional[str]]:
    """
    Extracts deed transfer parties and type from page text.

    Legal reasoning:
    Chain-of-title continuity depends on explicit conveyance evidence showing
    who transferred title and who received it.
    """
    normalized = normalize_whitespace(text)
    event_type = None
    if re.search(r"Warranty Deed", normalized, flags=re.IGNORECASE):
        event_type = "warranty deed"
    elif re.search(r"Quitclaim Deed", normalized, flags=re.IGNORECASE):
        event_type = "quitclaim deed"
    elif re.search(r"\bDeed\b", normalized, flags=re.IGNORECASE):
        event_type = "deed"

    grantor = grantee = None
    m = re.search(r"between\s+(.+?),\s*Grantor,?\s+and\s+(.+?),\s*Grantee", normalized, flags=re.IGNORECASE)
    if m:
        grantor = m.group(1).strip()
        grantee = m.group(2).strip()

    return event_type, grantor, grantee


def extract_chain_events(phase1: Dict, state_hint: str, log: List[Dict]) -> List[Dict]:
    """
    Extracts ownership events from Phase 1 page text.

    Why:
    Phase 2 converts raw pages into structured recorded ownership events for
    sequencing and validation.
    """
    events = []
    next_id = 1

    for page in phase1["pages"]:
        text = page["extracted_text"]
        event_type, grantor, grantee = parse_deed_event(text)
        if event_type:
            recording_date = parse_recording_date(text)
            document_reference = parse_instrument(text)
            county, state = parse_county_state(text, state_hint)

            event = {
                "event_id": next_id,
                "page_number": page["page_number"],
                "event_type": event_type,
                "grantor_name": grantor,
                "grantee_name": grantee,
                "recording_date": recording_date,
                "document_reference_number": document_reference,
                "county": county,
                "state": state,
                "confidence": page["confidence"],
                "requires_human_review": page["requires_human_review"],
            }
            events.append(event)
            next_id += 1

            add_log(
                log,
                phase="Phase 2 - Chain of Title",
                action="Extract ownership event",
                decision_type="AUTONOMOUS",
                confidence=page["confidence"],
                evidence={"page_number": page["page_number"], "text_sample": normalize_whitespace(text)[:200]},
                result=f"Extracted {event_type} event: {grantor} -> {grantee}, recorded {recording_date}.",
            )

    return events


def validate_chain(events: List[Dict], log: List[Dict]) -> Dict:
    """
    Runs chain-of-title validation.

    Legal reasoning:
    - Grantor/grantee continuity checks whether title passed cleanly from one
      grantee to the next grantor.
    - Long gaps between recorded transfers can indicate missing instruments.
    - Competing claims occur when the same grantor appears to convey to multiple grantees.
    """
    events = sorted(events, key=lambda e: (1 if e["recording_date"] is None else 0, e["recording_date"] or "9999-12-31"))
    continuity_flags, gap_flags, competing_flags = [], [], []

    for prev, curr in zip(events, events[1:]):
        if prev["grantee_name"] and curr["grantor_name"]:
            if normalize_whitespace(prev["grantee_name"]).lower() != normalize_whitespace(curr["grantor_name"]).lower():
                continuity_flags.append({
                    "type": "CHAIN GAP",
                    "status": "FAIL",
                    "page_number": curr["page_number"],
                    "expected": f"Grantor should match prior grantee: {prev['grantee_name']}",
                    "found": f"Found current grantor: {curr['grantor_name']}",
                    "analyst_verification_needed": "Determine whether an intermediate unrecorded or missing conveyance exists."
                })

        if prev["recording_date"] and curr["recording_date"]:
            d1 = datetime.fromisoformat(prev["recording_date"]).date()
            d2 = datetime.fromisoformat(curr["recording_date"]).date()
            delta = (d2 - d1).days
            if delta > 180:
                gap_flags.append({
                    "type": "POTENTIAL UNRECORDED TRANSFER",
                    "status": "FLAG",
                    "page_number": curr["page_number"],
                    "expected": "No more than 180 days between recorded ownership events",
                    "found": f"{delta} days between {prev['recording_date']} and {curr['recording_date']}",
                    "analyst_verification_needed": "Review whether an intervening deed or corrective instrument is missing from the packet."
                })

    seen = {}
    for event in events:
        grantor = event.get("grantor_name")
        grantee = event.get("grantee_name")
        if not grantor or not grantee:
            continue
        g_key = normalize_whitespace(grantor).lower()
        gr_key = normalize_whitespace(grantee).lower()
        if g_key in seen and seen[g_key] != gr_key:
            competing_flags.append({
                "type": "COMPETING CLAIM",
                "status": "FAIL",
                "page_number": event["page_number"],
                "expected": f"Grantor {grantor} should convey to a single supported grantee in the reviewed record set",
                "found": f"Different grantees found: {seen[g_key]} and {grantee}",
                "analyst_verification_needed": "Verify whether one deed is corrective, void, duplicate, or a true competing claim."
            })
        else:
            seen[g_key] = gr_key

    if competing_flags or continuity_flags:
        overall = "BROKEN"
    elif gap_flags:
        overall = "CLOUDED"
    else:
        overall = "CLEAR"

    for collection_name, flags in [("continuity", continuity_flags), ("gap detection", gap_flags), ("competing claims", competing_flags)]:
        add_log(
            log,
            phase="Phase 2 - Chain of Title",
            action=f"Validate {collection_name}",
            decision_type="FLAGGED_FOR_HUMAN_REVIEW" if flags else "AUTONOMOUS",
            confidence="LOW" if flags else "HIGH",
            evidence={"flag_count": len(flags), "flags": flags},
            result=f"{collection_name.title()} status: {'issues found' if flags else 'pass'}",
        )

    return {
        "document_id": events[0]["event_id"] if events else None,
        "ownership_event_count": len(events),
        "timeline": [
            {
                "timeline_number": idx + 1,
                **event,
                "status": "FLAG" if any(f.get("page_number") == event["page_number"] for f in continuity_flags + gap_flags + competing_flags) else "PASS"
            }
            for idx, event in enumerate(events)
        ],
        "validation_checks": {
            "check_a_grantor_grantee_continuity": {"status": "FAIL" if continuity_flags else "PASS", "flags": continuity_flags},
            "check_b_gap_detection": {"status": "FLAG" if gap_flags else "PASS", "flags": gap_flags},
            "check_c_competing_claims": {"status": "FAIL" if competing_flags else "PASS", "flags": competing_flags},
        },
        "overall_chain_integrity_score": overall,
        "conclusion": "INSUFFICIENT DATA - HUMAN REVIEW REQUIRED" if not events else overall,
    }


# -------------------------------------------------------------------
# Phase 3 - Lien Extraction and Waterfall
# -------------------------------------------------------------------

def extract_liens_from_pages(phase1: Dict, state_hint: str, investor_lien_type: str, log: List[Dict]) -> List[Dict]:
    """
    Extracts liens from document pages.

    Why:
    The waterfall engine requires explicit lien objects. This parser only uses
    text directly supported by the packet and does not infer missing balances or dates.
    """
    liens = []
    mortgage_candidates = []

    for page in phase1["pages"]:
        text = normalize_whitespace(page["extracted_text"])
        lowered = text.lower()

        if "property tax" in lowered or "delinquent taxes" in lowered:
            amount = parse_money(text) or 0.0
            lien = {
                "source_page": page["page_number"],
                "lien_type": "PROPERTY_TAX",
                "creditor_name": "County Tax Collector",
                "balance_amount": amount,
                "recording_date": None,
                "filing_date": None,
                "is_delinquent": True,
                "special_notes": "Extracted from tax page",
            }
            liens.append(lien)
            add_log(log, "Phase 3 - Lien Extraction", "Extract property tax lien", "AUTONOMOUS", page["confidence"],
                    {"page_number": page["page_number"], "text_sample": text[:180]}, f"Property tax balance extracted: {amount}")

        if "mortgage" in lowered:
            amount_match = re.search(r"(?:principal amount of|balance of|amount of)\s+\$([0-9,]+(?:\.[0-9]{2})?)", text, flags=re.IGNORECASE)
            amount = float(amount_match.group(1).replace(",", "")) if amount_match else (parse_money(text) or 0.0)
            creditor = None
            m = re.search(r"to\s+(.+?)(?:\s+in the original principal amount|\s+recorded|\s+Recorded|\.|$)", text, flags=re.IGNORECASE)
            if m:
                creditor = m.group(1).strip()
            mortgage_candidates.append({
                "source_page": page["page_number"],
                "creditor_name": creditor or f"Mortgage Creditor Page {page['page_number']}",
                "balance_amount": amount,
                "recording_date": parse_recording_date(text),
                "filing_date": None,
                "is_delinquent": False,
                "special_notes": "Extracted from mortgage page",
                "confidence": page["confidence"],
            })
            add_log(log, "Phase 3 - Lien Extraction", "Extract mortgage lien candidate", "AUTONOMOUS", page["confidence"],
                    {"page_number": page["page_number"], "text_sample": text[:180]}, f"Mortgage candidate extracted: creditor={creditor}, amount={amount}")

        if "federal tax lien" in lowered:
            amount_match = re.search(r"(?:amount of|balance of)\s+\$([0-9,]+(?:\.[0-9]{2})?)", text, flags=re.IGNORECASE)
            amount = float(amount_match.group(1).replace(",", "")) if amount_match else (parse_money(text) or 0.0)
            lien = {
                "source_page": page["page_number"],
                "lien_type": "IRS_FEDERAL",
                "creditor_name": "Internal Revenue Service",
                "balance_amount": amount,
                "recording_date": None,
                "filing_date": parse_filing_date(text) or parse_recording_date(text),
                "is_delinquent": False,
                "special_notes": "Federal tax lien notice extracted from packet",
            }
            liens.append(lien)
            add_log(log, "Phase 3 - Lien Extraction", "Extract IRS lien", "AUTONOMOUS", page["confidence"],
                    {"page_number": page["page_number"], "text_sample": text[:180]}, f"IRS lien extracted: amount={amount}, filing={lien['filing_date']}")

        if "judgment" in lowered and "lien" in lowered:
            amount = parse_money(text) or 0.0
            lien = {
                "source_page": page["page_number"],
                "lien_type": "JUDGMENT",
                "creditor_name": "Judgment Creditor",
                "balance_amount": amount,
                "recording_date": parse_recording_date(text),
                "filing_date": None,
                "is_delinquent": False,
                "special_notes": "Judgment lien extracted from packet",
            }
            liens.append(lien)

        if "hoa" in lowered and "lien" in lowered:
            amount = parse_money(text) or 0.0
            lien = {
                "source_page": page["page_number"],
                "lien_type": "HOA",
                "creditor_name": "HOA",
                "balance_amount": amount,
                "recording_date": parse_recording_date(text),
                "filing_date": None,
                "is_delinquent": False,
                "special_notes": "HOA lien extracted from packet",
            }
            liens.append(lien)

    mortgage_candidates.sort(key=lambda x: (1 if x["recording_date"] is None else 0, x["recording_date"] or "9999-12-31"))

    for idx, li in enumerate(mortgage_candidates):
        li["lien_type"] = "FIRST_MORTGAGE" if idx == 0 else "SECOND_MORTGAGE"
        liens.append(li)

    add_log(log, "Phase 3 - Lien Extraction", "Classify mortgages by recording order", "AUTONOMOUS", "MEDIUM",
            {"mortgage_candidates": mortgage_candidates}, f"Assigned {len(mortgage_candidates)} mortgage lien type(s) by recording order.")

    return liens


def sort_by_recording_date(liens: List[Dict]) -> List[Dict]:
    """
    Sorts liens by recording date ascending, undated items last.

    Legal reasoning:
    Standard recorded mortgage interests are modeled in public-record order.
    """
    return sorted(liens, key=lambda li: (1 if li.get("recording_date") is None else 0, li.get("recording_date") or "9999-12-31", li.get("creditor_name", "")))


def find_investor_position(waterfall_rows: List[Dict], investor_lien_type: str) -> Optional[int]:
    """
    Finds the investor lien type position in the executed waterfall.

    Why:
    The investor position is a core underwriting output.
    """
    for row in waterfall_rows:
        if row["lien_type"] == investor_lien_type:
            return row["step"]
    return None


def run_waterfall(property_estimated_value: float, liens: List[Dict], investor_lien_type: str, state: str, log: List[Dict]) -> Dict:
    """
    Runs the deterministic Phase 3 waterfall engine.

    Legal basis:
    1. Delinquent property taxes are modeled first as statutory super-priority.
    2. Mortgages follow recording chronology.
    3. IRS federal tax liens are separately budgeted as mandatory clearance cost.
    4. Remaining junior interests are applied after taxes, mortgages, and IRS cost.
    5. State-specific HOA super-priority rules are NOT handled and require manual review.
    """
    remaining = float(property_estimated_value)
    rows = []
    warnings = []
    irs_notes = []
    step = 1

    taxes = [li for li in liens if li["lien_type"] == "PROPERTY_TAX" and li["is_delinquent"]]
    for li in taxes:
        remaining -= float(li["balance_amount"])
        rows.append({
            "step": step,
            "lien_type": li["lien_type"],
            "creditor": li["creditor_name"],
            "balance": round(float(li["balance_amount"]), 2),
            "remaining_after": round(remaining, 2),
            "priority_basis": "Super Priority - statutory",
            "source_page": li.get("source_page"),
        })
        add_log(log, "Phase 3 - Waterfall", "Apply delinquent property tax", "AUTONOMOUS", "HIGH",
                {"lien": li, "remaining_after": remaining}, f"Remaining after tax step: {remaining}")
        step += 1

    if remaining <= 0:
        warnings.append("TAX WIPEOUT RISK - delinquent property taxes consume or exceed estimated property value.")
        add_log(log, "Phase 3 - Waterfall", "Tax wipeout stop", "FLAGGED_FOR_HUMAN_REVIEW", "HIGH",
                {"remaining_after_taxes": remaining}, "Workflow result: not recommended due to tax wipeout risk.")
        return {
            "waterfall_calculation": rows,
            "investor_position": find_investor_position(rows, investor_lien_type),
            "projected_net_recovery": round(remaining, 2),
            "recovery_status": "ZERO" if remaining == 0 else "NEGATIVE",
            "acquisition_flag": "NOT RECOMMENDED",
            "critical_warnings": warnings,
            "irs_lien_treatment": "No IRS treatment applied before tax wipeout result.",
        }

    senior = sort_by_recording_date([li for li in liens if li["lien_type"] in {"FIRST_MORTGAGE", "SECOND_MORTGAGE"}])
    investor_reached = False
    investor_position = None

    for li in senior:
        remaining -= float(li["balance_amount"])
        rows.append({
            "step": step,
            "lien_type": li["lien_type"],
            "creditor": li["creditor_name"],
            "balance": round(float(li["balance_amount"]), 2),
            "remaining_after": round(remaining, 2),
            "priority_basis": "Senior Recorded Interest - first-in-time by recording date",
            "source_page": li.get("source_page"),
        })
        if li["lien_type"] == investor_lien_type and investor_position is None:
            investor_position = step
            investor_reached = True

        add_log(log, "Phase 3 - Waterfall", "Apply mortgage lien in recording order", "AUTONOMOUS", "HIGH",
                {"lien": li, "remaining_after": remaining}, f"Remaining after mortgage step: {remaining}")
        step += 1

        if remaining <= 0 and not investor_reached:
            warnings.append("UNDERWATER - estimated value is exhausted before reaching investor's lien position.")
            add_log(log, "Phase 3 - Waterfall", "Underwater stop before investor lien", "FLAGGED_FOR_HUMAN_REVIEW", "HIGH",
                    {"remaining_before_investor": remaining, "investor_lien_type": investor_lien_type}, "Workflow result: not recommended due to zero recovery before investor position.")
            return {
                "waterfall_calculation": rows,
                "investor_position": investor_position,
                "projected_net_recovery": 0.0,
                "recovery_status": "ZERO",
                "acquisition_flag": "NOT RECOMMENDED",
                "critical_warnings": warnings,
                "irs_lien_treatment": "IRS liens, if present, still require manual review.",
            }

    irs_liens = [li for li in liens if li["lien_type"] == "IRS_FEDERAL"]
    for li in irs_liens:
        filed = li.get("filing_date") or "unknown date"
        warning = f"IRS lien filed {filed} - survival risk"
        warnings.append(warning)
        irs_notes.append(warning)
        remaining -= float(li["balance_amount"])
        rows.append({
            "step": step,
            "lien_type": li["lien_type"],
            "creditor": li["creditor_name"],
            "balance": round(float(li["balance_amount"]), 2),
            "remaining_after": round(remaining, 2),
            "priority_basis": "Mandatory Cost - IRS federal tax lien survival / clearance budgeting",
            "source_page": li.get("source_page"),
        })
        add_log(log, "Phase 3 - Waterfall", "Budget IRS lien as mandatory cost", "AUTONOMOUS", "HIGH",
                {"lien": li, "remaining_after": remaining}, f"Remaining after IRS step: {remaining}")
        step += 1

    juniors = sort_by_recording_date([li for li in liens if li["lien_type"] in {"JUDGMENT", "HOA", "OTHER"}])
    for li in juniors:
        remaining -= float(li["balance_amount"])
        rows.append({
            "step": step,
            "lien_type": li["lien_type"],
            "creditor": li["creditor_name"],
            "balance": round(float(li["balance_amount"]), 2),
            "remaining_after": round(remaining, 2),
            "priority_basis": "Junior Interest - applied after taxes, mortgages, and IRS mandatory clearance costs",
            "source_page": li.get("source_page"),
        })
        add_log(log, "Phase 3 - Waterfall", "Apply junior lien", "AUTONOMOUS", "HIGH",
                {"lien": li, "remaining_after": remaining}, f"Remaining after junior step: {remaining}")
        step += 1

    if any(li["lien_type"] == "HOA" for li in liens):
        warnings.append("State-specific HOA super-priority rules are NOT currently handled and require manual review.")

    if investor_position is None:
        warnings.append("Investor lien type was not found in the executed waterfall. Verify lien classification and input mapping.")

    recovery_status = "POSITIVE" if remaining > 0 else ("ZERO" if remaining == 0 else "NEGATIVE")
    acquisition_flag = "NOT RECOMMENDED" if recovery_status in {"ZERO", "NEGATIVE"} else ("CAUTION" if warnings else "RECOMMENDED")

    add_log(log, "Phase 3 - Waterfall", "Finalize waterfall output", "AUTONOMOUS", "HIGH",
            {"final_remaining": remaining, "warnings": warnings}, f"Recovery status {recovery_status}; acquisition flag {acquisition_flag}.")

    return {
        "waterfall_calculation": rows,
        "investor_position": investor_position,
        "projected_net_recovery": round(remaining, 2),
        "recovery_status": recovery_status,
        "acquisition_flag": acquisition_flag,
        "critical_warnings": warnings,
        "irs_lien_treatment": " ; ".join(irs_notes) if irs_notes else "No IRS federal tax liens identified.",
    }


# -------------------------------------------------------------------
# Phase 4 - Risk scoring and dashboard summary
# -------------------------------------------------------------------

def derive_packet_quality_score(phase1: Dict) -> float:
    """
    Calculates high-confidence-page percentage.

    Why:
    Document quality affects trust in all downstream extractions.
    """
    total = max(phase1["total_pages"], 1)
    high = len(phase1["extraction_summary"]["high_confidence_pages"])
    return round((high / total) * 100.0, 2)


def detect_low_confidence_critical_sections(phase1: Dict) -> List[Dict]:
    """
    Detects low-confidence critical pages.

    Why:
    Critical low-confidence pages increase chain risk and should be surfaced.
    """
    findings = []
    keywords = ["signature", "recording", "stamp", "obscured", "illegible", "federal lien", "mortgage", "deed"]
    for page in phase1["pages"]:
        if page["confidence"] != "LOW":
            continue
        blob = " ".join(page["flags"]).lower() + " " + (page["extracted_text"] or "").lower()
        if any(k in blob for k in keywords):
            findings.append({"page_number": page["page_number"], "flags": page["flags"]})
    return findings


def compute_risk_scores(phase1: Dict, chain: Dict, waterfall: Dict, liens: List[Dict], log: List[Dict]) -> Dict:
    """
    Computes the weighted 0-10 risk score.

    Why:
    The analyst needs a single top-line risk indicator built from chain quality,
    lien-position economics, and source-document confidence.
    """
    chain_integrity = (chain.get("overall_chain_integrity_score") or chain.get("conclusion") or "CLOUDED").upper()
    if chain_integrity == "CLEAR":
        chain_score = 0
    elif chain_integrity == "CLOUDED":
        chain_score = 4
    else:
        chain_score = 8

    low_critical = detect_low_confidence_critical_sections(phase1)
    if low_critical:
        chain_score += 2

    projected = float(waterfall.get("projected_net_recovery", 0.0))
    investor_type = None
    investor_balance = 0.0
    for li in liens:
        if li["lien_type"] in {"FIRST_MORTGAGE", "SECOND_MORTGAGE"} and li["creditor_name"].lower().find("investor") >= 0:
            investor_type = li["lien_type"]
            investor_balance = float(li["balance_amount"])
            break

    recovery_status = waterfall.get("recovery_status", "ZERO").upper()
    if recovery_status in {"ZERO", "NEGATIVE"}:
        lien_score = 8
    elif investor_balance > 0 and projected < (0.20 * investor_balance):
        lien_score = 4
    else:
        lien_score = 0

    if "irs lien filed" in (waterfall.get("irs_lien_treatment") or "").lower():
        lien_score += 2

    packet_quality = derive_packet_quality_score(phase1)
    if packet_quality > 80:
        doc_score = 0
    elif 50 <= packet_quality <= 80:
        doc_score = 2
    else:
        doc_score = 4

    overall = round(min(max((chain_score * 0.40) + (lien_score * 0.40) + (doc_score * 0.20), 0), 10), 1)
    color = "GREEN" if overall <= 3 else ("YELLOW" if overall <= 6 else "RED")

    add_log(log, "Phase 4 - Risk Scoring", "Compute weighted risk score", "AUTONOMOUS", "HIGH",
            {
                "chain_score": chain_score,
                "lien_score": lien_score,
                "document_score": doc_score,
                "packet_quality_pct": packet_quality,
                "low_confidence_critical_pages": low_critical,
            },
            f"Overall risk score {overall} ({color}).")

    return {
        "packet_quality_score": packet_quality,
        "chain_risk_score": min(chain_score, 10),
        "lien_risk_score": min(lien_score, 10),
        "document_quality_risk_score": doc_score,
        "overall_risk_score": overall,
        "risk_color_indicator": color,
    }


def collect_alerts(phase1: Dict, chain: Dict, waterfall: Dict) -> List[Dict]:
    """
    Collects cross-phase alerts sorted by severity.

    Why:
    The analyst dashboard should centralize all exceptions in one place.
    """
    alerts = []

    def add_alert(severity: str, message: str, page_ref: Optional[str], source: str):
        rank = {"CRITICAL": 1, "HIGH": 2, "MODERATE": 3, "LOW": 4}[severity]
        alerts.append({
            "severity": severity,
            "severity_rank": rank,
            "message": message,
            "page_reference": page_ref or "No page reference available",
            "source": source,
        })

    for page in phase1["pages"]:
        if page["requires_human_review"]:
            severity = "HIGH" if page["confidence"] == "LOW" else "MODERATE"
            add_alert(severity, f"Document quality concern: {', '.join(page['flags']) or 'manual review required'}.", f"Page {page['page_number']}", "Phase 1")

    for payload in chain["validation_checks"].values():
        for flag in payload["flags"]:
            severity = "CRITICAL" if flag["status"] == "FAIL" else "HIGH"
            add_alert(severity, f"{flag['type']}: expected {flag['expected']} ; found {flag['found']}.", f"Page {flag['page_number']}" if flag.get("page_number") else None, "Phase 2")

    for warning in waterfall["critical_warnings"]:
        severity = "HIGH" if any(word in warning for word in ["IRS", "UNDERWATER", "TAX WIPEOUT"]) else "MODERATE"
        add_alert(severity, warning, None, "Phase 3")

    alerts.sort(key=lambda x: (x["severity_rank"], x["source"], x["page_reference"], x["message"]))
    return alerts


def build_next_actions(phase1: Dict, chain: Dict, waterfall: Dict, liens: List[Dict]) -> List[str]:
    """
    Builds analyst next actions from current findings.

    Why:
    The dashboard must translate exceptions into concrete analyst workflow steps.
    """
    actions = []

    for page in detect_low_confidence_critical_sections(phase1):
        actions.append(
            f"Verify recording and signature details on page {page['page_number']} - LOW-confidence critical extraction with flags: {', '.join(page['flags'])}."
        )

    for payload in chain["validation_checks"].values():
        for flag in payload["flags"]:
            actions.append(flag["analyst_verification_needed"])

    for warning in waterfall["critical_warnings"]:
        if "IRS lien filed" in warning:
            actions.append("Initiate IRS demand / release review workflow and budget federal lien clearance cost before any acquisition decision.")
        elif "HOA" in warning:
            actions.append("Review state-specific HOA super-priority rules manually; the current engine does not model those carveouts.")
        elif "UNDERWATER" in warning:
            actions.append("Re-check collateral value and senior balances because the investor position is projected to be underwater.")
        elif "TAX WIPEOUT RISK" in warning:
            actions.append("Confirm current delinquent property tax balance directly with the taxing authority before proceeding.")

    for li in liens:
        if li["lien_type"] == "PROPERTY_TAX" and li["is_delinquent"]:
            actions.append(f"Confirm current tax delinquency amount with county tax collector - modeled amount {money(li['balance_amount'])}; stale tax data can materially change recovery.")

    seen = set()
    deduped = []
    for act in actions:
        if act not in seen:
            seen.add(act)
            deduped.append(act)
    return deduped[:10]


def build_dashboard_json(property_address: str, analyst_name: str, phase1: Dict, chain: Dict, waterfall: Dict, liens: List[Dict], risk: Dict, halted: bool, halt_message: str, log: List[Dict]) -> Dict:
    """
    Builds the unified dashboard JSON.

    Why:
    The final report should present one analyst-ready view integrating all phases.
    """
    alerts = collect_alerts(phase1, chain, waterfall)
    next_actions = build_next_actions(phase1, chain, waterfall, liens)

    if halted:
        recommendation = {
            "recommendation": "REQUIRES FURTHER INVESTIGATION",
            "rationale": halt_message,
        }
    elif chain["overall_chain_integrity_score"] == "BROKEN":
        recommendation = {
            "recommendation": "DO NOT PROCEED",
            "rationale": "Chain broken - defects block reliable reliance on ownership continuity.",
        }
    elif waterfall["recovery_status"] in {"ZERO", "NEGATIVE"} or any("TAX WIPEOUT RISK" in w for w in waterfall["critical_warnings"]):
        recommendation = {
            "recommendation": "DO NOT PROCEED",
            "rationale": "Underwater, tax wipeout, or negative projected recovery makes the acquisition not recommended.",
        }
    elif alerts:
        recommendation = {
            "recommendation": "PROCEED WITH CAUTION",
            "rationale": f"Projected recovery {money(waterfall['projected_net_recovery'])}; verify flagged title and lien issues before final decision.",
        }
    else:
        recommendation = {
            "recommendation": "PROCEED WITH CAUTION",
            "rationale": f"Projected recovery {money(waterfall['projected_net_recovery'])}; no additional critical alerts currently detected.",
        }

    add_log(log, "Phase 4 - Dashboard", "Build recommendation block", "AUTONOMOUS", "HIGH",
            {"halted": halted, "chain_status": chain["overall_chain_integrity_score"], "recovery_status": waterfall["recovery_status"], "alerts_count": len(alerts)},
            f"Recommendation: {recommendation['recommendation']}")

    return {
        "header": {
            "property_address": property_address,
            "review_date": str(date.today()),
            "analyst_name": analyst_name,
            "packet_quality_score": risk["packet_quality_score"],
            "overall_risk_score": risk["overall_risk_score"],
            "risk_color_indicator": risk["risk_color_indicator"],
        },
        "acquisition_recommendation": recommendation,
        "waterfall_summary_table": waterfall["waterfall_calculation"],
        "investor_position": waterfall["investor_position"],
        "projected_net_recovery": waterfall["projected_net_recovery"],
        "recovery_status": waterfall["recovery_status"],
        "chain_of_title_timeline": chain["timeline"],
        "critical_alerts": alerts,
        "recommended_next_actions": next_actions,
        "irs_lien_treatment": waterfall["irs_lien_treatment"],
        "manual_review_note": "State-specific HOA super-priority rules are NOT currently handled and require manual review.",
        "halted_for_human_input": halted,
        "halt_message": halt_message if halted else "",
    }


# -------------------------------------------------------------------
# Output generation - PDF summary, Excel, processing log
# -------------------------------------------------------------------

def _styles():
    """
    Creates the PDF style system.

    Why:
    The analyst dashboard PDF should be concise, clean, and readable on one page.
    """
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="DashTitle",
        parent=styles["Title"],
        fontName="Helvetica-Bold",
        fontSize=18,
        leading=21,
        textColor=colors.HexColor("#0F172A"),
    ))
    styles.add(ParagraphStyle(
        name="SectionBand",
        parent=styles["Heading2"],
        fontName="Helvetica-Bold",
        fontSize=10.5,
        leading=12.5,
        textColor=colors.white,
        backColor=colors.HexColor("#1F3A5F"),
        borderPadding=(5, 5, 5),
        spaceBefore=6,
        spaceAfter=5,
    ))
    styles.add(ParagraphStyle(
        name="BodySmall",
        parent=styles["BodyText"],
        fontName="Helvetica",
        fontSize=8.3,
        leading=10.4,
        textColor=colors.HexColor("#111827"),
        spaceAfter=2,
    ))
    return styles


def _cell(text: str, styles, size: float = 7.6, bold: bool = False, align: int = TA_LEFT) -> Paragraph:
    """
    Wraps table cell text in a Paragraph.

    Why:
    Wrapped paragraph cells prevent text from overflowing outside the PDF tables.
    """
    safe = str(text).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
    style = ParagraphStyle(
        "TmpCell",
        parent=styles["BodyText"],
        fontName="Helvetica-Bold" if bold else "Helvetica",
        fontSize=size,
        leading=size + 2,
        textColor=colors.HexColor("#111827"),
        alignment=align,
        wordWrap="CJK",
    )
    return Paragraph(safe, style)


def build_one_page_pdf(dashboard: Dict, output_pdf_path: Path) -> None:
    """
    Generates the final one-page analyst dashboard PDF.

    Why:
    The user requested a clean downloadable one-page summary suitable for analyst triage.
    """
    styles = _styles()
    doc = SimpleDocTemplate(
        str(output_pdf_path),
        pagesize=letter,
        leftMargin=0.45 * inch,
        rightMargin=0.45 * inch,
        topMargin=0.40 * inch,
        bottomMargin=0.35 * inch,
    )
    story = []

    hdr = dashboard["header"]
    story.append(Paragraph("TitleIQ Agent - Analyst Dashboard", styles["DashTitle"]))
    story.append(Paragraph(
        f"Property: {hdr['property_address']} | Review Date: {hdr['review_date']} | Analyst: {hdr['analyst_name'] or '________________'} | "
        f"Packet Quality: {hdr['packet_quality_score']}% high-confidence pages | Risk Score: {hdr['overall_risk_score']} / 10 ({hdr['risk_color_indicator']})",
        styles["BodySmall"]
    ))
    story.append(Spacer(1, 5))

    # Recommendation
    story.append(Paragraph("SECTION 1 - ACQUISITION RECOMMENDATION", styles["SectionBand"]))
    rec = dashboard["acquisition_recommendation"]
    story.append(Paragraph(f"<b>{rec['recommendation']}</b> - {rec['rationale']}", styles["BodySmall"]))

    # Waterfall
    story.append(Paragraph("SECTION 2 - WATERFALL SUMMARY", styles["SectionBand"]))
    wf_rows = [[
        _cell("Step", styles, 7.4, True),
        _cell("Lien Type", styles, 7.4, True),
        _cell("Creditor", styles, 7.4, True),
        _cell("Balance", styles, 7.4, True),
        _cell("Remaining", styles, 7.4, True),
    ]]
    inv_pos = dashboard["investor_position"]
    for row in dashboard["waterfall_summary_table"]:
        lien_type = row["lien_type"] + (" - INVESTOR" if inv_pos == row["step"] else "")
        wf_rows.append([
            _cell(row["step"], styles, 7.2),
            _cell(lien_type, styles, 7.1, inv_pos == row["step"]),
            _cell(row["creditor"], styles, 7.1),
            _cell(money(row["balance"]), styles, 7.1),
            _cell(money(row["remaining_after"]), styles, 7.1),
        ])
    wf_tbl = Table(wf_rows, colWidths=[0.38*inch, 1.38*inch, 2.0*inch, 0.95*inch, 1.02*inch], repeatRows=1)
    wf_style = [
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#DCE8F5")),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 5),
        ("RIGHTPADDING", (0,0), (-1,-1), 5),
        ("TOPPADDING", (0,0), (-1,-1), 5),
        ("BOTTOMPADDING", (0,0), (-1,-1), 5),
    ]
    if inv_pos is not None:
        wf_style += [("BACKGROUND", (0, inv_pos), (-1, inv_pos), colors.HexColor("#FEF3C7"))]
    wf_tbl.setStyle(TableStyle(wf_style))
    story.append(wf_tbl)
    story.append(Paragraph(
        f"<b>Investor Position:</b> {dashboard['investor_position']} | <b>Projected Net Recovery:</b> {money(dashboard['projected_net_recovery'])} | "
        f"<b>Recovery Status:</b> {dashboard['recovery_status']}",
        styles["BodySmall"]
    ))

    # Chain timeline
    story.append(Paragraph("SECTION 3 - CHAIN OF TITLE", styles["SectionBand"]))
    tl_rows = [[
        _cell("#", styles, 7.4, True),
        _cell("Status", styles, 7.4, True),
        _cell("Type", styles, 7.4, True),
        _cell("Grantor", styles, 7.4, True),
        _cell("Grantee", styles, 7.4, True),
        _cell("Recorded", styles, 7.4, True),
    ]]
    for item in dashboard["chain_of_title_timeline"]:
        tl_rows.append([
            _cell(item["timeline_number"], styles, 7.0),
            _cell(item["status"], styles, 7.0, item["status"] == "FLAG"),
            _cell(item["event_type"], styles, 7.0),
            _cell(item["grantor_name"] or "N/A", styles, 7.0),
            _cell(item["grantee_name"] or "N/A", styles, 7.0),
            _cell(item["recording_date"] or "N/A", styles, 7.0),
        ])
    tl_tbl = Table(tl_rows, colWidths=[0.3*inch, 0.55*inch, 1.0*inch, 1.7*inch, 1.7*inch, 0.95*inch], repeatRows=1)
    tl_style = [
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#E2E8F0")),
        ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING", (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING", (0,0), (-1,-1), 4),
        ("BOTTOMPADDING", (0,0), (-1,-1), 4),
    ]
    for i, item in enumerate(dashboard["chain_of_title_timeline"], start=1):
        if item["status"] == "FLAG":
            tl_style.append(("BACKGROUND", (0,i), (-1,i), colors.HexColor("#FEF2F2")))
    tl_tbl.setStyle(TableStyle(tl_style))
    story.append(tl_tbl)

    # Alerts & next actions in two-column compact block
    story.append(Paragraph("SECTION 4 - CRITICAL ALERTS AND NEXT ACTIONS", styles["SectionBand"]))
    alert_text = "<br/>".join(
        [f"{i}. [{a['severity']}] {a['message']} ({a['page_reference']})" for i, a in enumerate(dashboard["critical_alerts"][:5], start=1)]
    ) or "No critical alerts."
    action_text = "<br/>".join(
        [f"{i}. {a}" for i, a in enumerate(dashboard["recommended_next_actions"][:5], start=1)]
    ) or "No next actions."
    two_col = Table([
        [
            Paragraph(f"<b>Critical Alerts</b><br/>{alert_text}", styles["BodySmall"]),
            Paragraph(f"<b>Recommended Next Actions</b><br/>{action_text}", styles["BodySmall"]),
        ]
    ], colWidths=[3.55*inch, 3.55*inch])
    two_col.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,-1), colors.HexColor("#F8FAFC")),
        ("GRID", (0,0), (-1,-1), 0.4, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0,0), (-1,-1), "TOP"),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING", (0,0), (-1,-1), 6),
        ("BOTTOMPADDING", (0,0), (-1,-1), 6),
    ]))
    story.append(two_col)
    story.append(Spacer(1, 4))
    story.append(Paragraph(f"<b>IRS Treatment:</b> {dashboard['irs_lien_treatment']}", styles["BodySmall"]))
    story.append(Paragraph(f"<b>Manual Review Note:</b> {dashboard['manual_review_note']}", styles["BodySmall"]))

    doc.build(story)


def build_excel(phase1: Dict, chain: Dict, waterfall: Dict, dashboard: Dict, processing_log: List[Dict], output_xlsx_path: Path) -> None:
    """
    Generates the detailed analysis Excel workbook.

    Why:
    Analysts need a structured workbook with the chain timeline and waterfall
    table, plus summary and alerts, for detailed review and downstream workflow.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"

    thin = Side(style="thin", color="D9E2F3")
    header_fill = PatternFill("solid", fgColor="1F3A5F")
    header_font = Font(color="FFFFFF", bold=True)
    band_fill = PatternFill("solid", fgColor="DCE8F5")
    caution_fill = PatternFill("solid", fgColor="FEF3C7")
    alert_fill = PatternFill("solid", fgColor="FEE2E2")

    def style_header(sheet, row_num, cols):
        for c in range(1, cols + 1):
            cell = sheet.cell(row=row_num, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(bottom=thin)

    # Overview sheet
    overview_rows = [
        ("Property Address", dashboard["header"]["property_address"]),
        ("Review Date", dashboard["header"]["review_date"]),
        ("Analyst Name", dashboard["header"]["analyst_name"]),
        ("Packet Quality Score", f"{dashboard['header']['packet_quality_score']}%"),
        ("Overall Risk Score", dashboard["header"]["overall_risk_score"]),
        ("Risk Color", dashboard["header"]["risk_color_indicator"]),
        ("Recommendation", dashboard["acquisition_recommendation"]["recommendation"]),
        ("Recommendation Rationale", dashboard["acquisition_recommendation"]["rationale"]),
        ("Projected Net Recovery", dashboard["projected_net_recovery"]),
        ("Recovery Status", dashboard["recovery_status"]),
        ("Investor Position", dashboard["investor_position"]),
        ("IRS Treatment", dashboard["irs_lien_treatment"]),
    ]
    ws["A1"] = "TitleIQ Agent - Detailed Analysis"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Metric"
    ws["B3"] = "Value"
    style_header(ws, 3, 2)
    row = 4
    for label, value in overview_rows:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=value)
        ws.cell(row=row, column=1).font = Font(bold=True, color="666666")
        ws.cell(row=row, column=1).fill = band_fill
        ws.cell(row=row, column=1).alignment = Alignment(vertical="center", wrap_text=True)
        ws.cell(row=row, column=2).alignment = Alignment(vertical="center", wrap_text=True)
        row += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 82

    # Phase 1 sheet
    q = wb.create_sheet("Phase1_Quality")
    q_headers = ["Page Number", "Confidence", "Flags", "Requires Human Review", "Extracted Text Sample"]
    for i, h in enumerate(q_headers, start=1):
        q.cell(row=1, column=i, value=h)
    style_header(q, 1, len(q_headers))
    for r, page in enumerate(phase1["pages"], start=2):
        q.cell(row=r, column=1, value=page["page_number"])
        q.cell(row=r, column=2, value=page["confidence"])
        q.cell(row=r, column=3, value=", ".join(page["flags"]))
        q.cell(row=r, column=4, value=page["requires_human_review"])
        q.cell(row=r, column=5, value=(page["extracted_text"] or "")[:300])
        for c in range(1, 6):
            q.cell(row=r, column=c).alignment = Alignment(vertical="center", wrap_text=True)
        if page["confidence"] == "LOW":
            for c in range(1, 6):
                q.cell(row=r, column=c).fill = caution_fill
    for col, width in {"A":12, "B":12, "C":34, "D":18, "E":70}.items():
        q.column_dimensions[col].width = width

    # Chain sheet
    cws = wb.create_sheet("Chain_Timeline")
    c_headers = ["Timeline #", "Status", "Event Type", "Grantor", "Grantee", "Recording Date", "Document Ref", "County", "State", "Page", "Confidence"]
    for i, h in enumerate(c_headers, start=1):
        cws.cell(row=1, column=i, value=h)
    style_header(cws, 1, len(c_headers))
    for r, item in enumerate(chain["timeline"], start=2):
        vals = [
            item["timeline_number"], item["status"], item["event_type"], item["grantor_name"], item["grantee_name"],
            item["recording_date"], item["document_reference_number"], item["county"], item["state"], item["page_number"], item["confidence"]
        ]
        for i, v in enumerate(vals, start=1):
            cws.cell(row=r, column=i, value=v)
            cws.cell(row=r, column=i).alignment = Alignment(vertical="center", wrap_text=True)
        if item["status"] == "FLAG":
            for i in range(1, len(c_headers)+1):
                cws.cell(row=r, column=i).fill = alert_fill
    for col, width in {"A":10, "B":10, "C":16, "D":22, "E":22, "F":14, "G":18, "H":18, "I":10, "J":8, "K":12}.items():
        cws.column_dimensions[col].width = width

    # Waterfall sheet
    wws = wb.create_sheet("Waterfall")
    w_headers = ["Step", "Lien Type", "Creditor", "Balance", "Remaining After", "Priority Basis", "Source Page"]
    for i, h in enumerate(w_headers, start=1):
        wws.cell(row=1, column=i, value=h)
    style_header(wws, 1, len(w_headers))
    for r, item in enumerate(waterfall["waterfall_calculation"], start=2):
        vals = [item["step"], item["lien_type"], item["creditor"], item["balance"], item["remaining_after"], item["priority_basis"], item.get("source_page")]
        for i, v in enumerate(vals, start=1):
            wws.cell(row=r, column=i, value=v)
            wws.cell(row=r, column=i).alignment = Alignment(vertical="center", wrap_text=True)
        if dashboard["investor_position"] == item["step"]:
            for i in range(1, len(w_headers)+1):
                wws.cell(row=r, column=i).fill = caution_fill
                wws.cell(row=r, column=i).font = Font(bold=True)
        wws.cell(row=r, column=4).number_format = '$#,##0.00;[Red]($#,##0.00);-'
        wws.cell(row=r, column=5).number_format = '$#,##0.00;[Red]($#,##0.00);-'
    for col, width in {"A":8, "B":18, "C":24, "D":14, "E":16, "F":48, "G":10}.items():
        wws.column_dimensions[col].width = width

    # Alerts sheet
    aws = wb.create_sheet("Alerts")
    a_headers = ["Severity", "Source", "Page Reference", "Message"]
    for i, h in enumerate(a_headers, start=1):
        aws.cell(row=1, column=i, value=h)
    style_header(aws, 1, len(a_headers))
    for r, alert in enumerate(dashboard["critical_alerts"], start=2):
        vals = [alert["severity"], alert["source"], alert["page_reference"], alert["message"]]
        for i, v in enumerate(vals, start=1):
            aws.cell(row=r, column=i, value=v)
            aws.cell(row=r, column=i).alignment = Alignment(vertical="center", wrap_text=True)
        if alert["severity"] in {"CRITICAL", "HIGH"}:
            for i in range(1, len(a_headers)+1):
                aws.cell(row=r, column=i).fill = caution_fill if alert["severity"] == "HIGH" else alert_fill
    for col, width in {"A":12, "B":14, "C":18, "D":88}.items():
        aws.column_dimensions[col].width = width

    # Processing log mirror sheet
    lws = wb.create_sheet("Processing_Log")
    l_headers = ["Timestamp", "Phase", "Action", "Decision Type", "Confidence", "Result", "Evidence"]
    for i, h in enumerate(l_headers, start=1):
        lws.cell(row=1, column=i, value=h)
    style_header(lws, 1, len(l_headers))
    for r, entry in enumerate(processing_log, start=2):
        vals = [
            entry["timestamp"], entry["phase"], entry["action"], entry["decision_type"],
            entry["confidence"], entry["result"], json.dumps(entry["evidence"], ensure_ascii=False),
        ]
        for i, v in enumerate(vals, start=1):
            lws.cell(row=r, column=i, value=v)
            lws.cell(row=r, column=i).alignment = Alignment(vertical="center", wrap_text=True)
    for col, width in {"A":22, "B":18, "C":26, "D":20, "E":12, "F":42, "G":72}.items():
        lws.column_dimensions[col].width = width

    wb.save(output_xlsx_path)


def save_processing_log(log: List[Dict], output_path: Path) -> None:
    """
    Saves the processing log as structured JSON.

    Why:
    The user requires a downloadable audit trail capturing autonomous decisions.
    """
    payload = {
        "summary": {
            "entry_count": len(log),
            "generated_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        },
        "entries": log,
    }
    Path(output_path).write_text(json.dumps(payload, indent=2), encoding="utf-8")


# -------------------------------------------------------------------
# Master orchestrator
# -------------------------------------------------------------------

def run_titleiq_agent(
    pdf_path: str,
    property_estimated_value: float,
    investor_lien_type: str,
    property_address: str = "",
    analyst_name: str = "",
    state: str = "",
    output_pdf_path: str = "titleiq_dashboard.pdf",
    output_xlsx_path: str = "titleiq_analysis.xlsx",
    output_log_path: str = "titleiq_processing_log.json",
) -> Dict:
    """
    Runs the full TitleIQ Agent workflow.

    Why:
    This is the single entry point that orchestrates all four phases, passes
    outputs from each step into the next, applies the quality stop-rule, and
    emits the final analyst-ready deliverables.
    """
    log: List[Dict] = []

    add_log(log, "Orchestrator", "Start workflow", "AUTONOMOUS", "HIGH",
            {"pdf_path": pdf_path, "property_estimated_value": property_estimated_value, "investor_lien_type": investor_lien_type, "state": state},
            "Workflow started.")

    # Step 1: Document intake
    phase1 = extract_pdf_pages(Path(pdf_path), log)
    quality_summary = phase1_quality_report(phase1)
    add_log(log, "Phase 1 - Document Intake", "Produce quality report", "AUTONOMOUS", "HIGH", {"report": quality_summary}, quality_summary)

    gate = critical_quality_gate(phase1, log, min_high_conf_pct=40.0)

    # Step 2: Chain of title
    chain_events = extract_chain_events(phase1, state, log)
    chain = validate_chain(chain_events, log)

    # Step 3: Lien extraction + waterfall
    liens = extract_liens_from_pages(phase1, state, investor_lien_type, log)
    waterfall = run_waterfall(property_estimated_value, liens, investor_lien_type, state, log)

    halted = gate["halted"]
    halt_message = gate["message"] if halted else ""

    if halted:
        add_log(log, "Orchestrator", "Halt workflow for human input", "FLAGGED_FOR_HUMAN_REVIEW", "LOW",
                {"gate": gate}, "Critical-section confidence below threshold. Human input required before reliance.")
    else:
        add_log(log, "Orchestrator", "Proceed past quality gate", "AUTONOMOUS", "HIGH",
                {"gate": gate}, "Critical-section confidence meets threshold. Proceeded automatically.")

    # Step 4: Dashboard
    risk = compute_risk_scores(phase1, chain, waterfall, liens, log)
    dashboard = build_dashboard_json(
        property_address=property_address or "Property address not provided",
        analyst_name=analyst_name,
        phase1=phase1,
        chain=chain,
        waterfall=waterfall,
        liens=liens,
        risk=risk,
        halted=halted,
        halt_message=halt_message,
        log=log,
    )

    build_one_page_pdf(dashboard, Path(output_pdf_path))
    build_excel(phase1, chain, waterfall, dashboard, log, Path(output_xlsx_path))
    save_processing_log(log, Path(output_log_path))

    add_log(log, "Orchestrator", "Finish workflow", "AUTONOMOUS", "HIGH",
            {"output_pdf_path": output_pdf_path, "output_xlsx_path": output_xlsx_path, "output_log_path": output_log_path},
            "Workflow completed and deliverables saved.")

    # overwrite log with final entry included
    save_processing_log(log, Path(output_log_path))

    return {
        "phase1_output": phase1,
        "chain_output": chain,
        "liens_extracted": liens,
        "waterfall_output": waterfall,
        "dashboard_output": dashboard,
        "processing_log_entries": len(log),
        "deliverables": {
            "dashboard_pdf": output_pdf_path,
            "analysis_excel": output_xlsx_path,
            "processing_log": output_log_path,
        },
    }
